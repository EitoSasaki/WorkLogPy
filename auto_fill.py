# outputディレクトリにある勤務ログファイルをもとにエクセルファイルの勤務表に自動入力する
# --dateオプションで指定された年月(YYYYMM)の勤務ログを対象とする
# --fileオプションで指定されたエクセルファイルに入力する
# 例: python auto_fill.py --date 202406 --file 勤務表.xlsx
# 勤務ログファイルは "output/YYYYMM_work_log.json" に保存されていると想定
# 勤務ログファイルの形式は work_log.py に準拠

import json
import os
import argparse
from datetime import datetime
import pathlib
from openpyxl import load_workbook

def load_work_log(file_path):
    # 勤務ログを読み込む
    if not os.path.exists(file_path):
        print(f"勤務ログファイルが見つかりません: {file_path}")
        return {}
    with open(file_path, 'r') as f:
        return json.load(f)
    
def fill_work_log_to_excel(log, excel_file, year_month):
    # エクセルファイルに勤務ログを自動入力する
    # 対象のシートは「YYYY年MM月」シートを想定
    wb = load_workbook(excel_file)
    
    year = int(year_month[:4])
    month = int(year_month[4:6])
    
    # シート名を "YYYY年MM月" として切り替え（なければアクティブシートを使用）
    sheet_name = f"{year}年{month:02d}月"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        print(f"警告: シート '{sheet_name}' が見つかりません。アクティブシートを使用します。")
        ws = wb.active

    # 日付と一致する行に勤務ログを入力する
    # 1日から31日までループ
    # 日付が存在しない場合はスキップ
    # C列: 勤務開始(時)、D列: 勤務開始(分)、F列: 勤務終了(時)、G列: 勤務終了(分)、J列: 休憩時間(1固定)
    # 日付はA列11行目から始まると仮定
    # entryが存在しない場合は空で埋める
    # 時、分は2桁で、1桁の場合は先頭に0を付ける
    for day in range(1, 32):
        try:
            date_obj = datetime(year, month, day)
        except ValueError:
            continue  # 無効な日付はスキップ

        date_str = date_obj.strftime("%Y-%m-%d")
        if date_str in log and isinstance(log[date_str], list) and len(log[date_str]) > 0:
            entry = log[date_str][0]
            # A列の11行目が1日なので、行は day + 10
            row = day + 10

            # 時刻文字列 "HH:MM" や数値に対応して (hour, minute) を返す
            def parse_time(val):
                if val is None:
                    return (None, None)
                if isinstance(val, str):
                    if ':' in val:
                        parts = val.split(':', 1)
                        try:
                            return int(parts[0]), int(parts[1])
                        except ValueError:
                            return (None, None)
                    else:
                        try:
                            return int(val), 0
                        except ValueError:
                            return (None, None)
                if isinstance(val, (int, float)):
                    return int(val), 0
                return (None, None)

            # 開始時刻 -> C列(3)=時、D列(4)=分
            if 'start' in entry:
                sh, sm = parse_time(entry['start'])
                if sh is not None:
                    sh = f"{sh:02d}"
                    ws.cell(row=row, column=3, value=sh)
                if sm is not None:
                    sm = f"{sm:02d}"
                    ws.cell(row=row, column=4, value=sm)
            else:
                ws.cell(row=row, column=3, value="")
                ws.cell(row=row, column=4, value="None")

            # 終了時刻 -> F列(6)=時、G列(7)=分
            if 'end' in entry:
                eh, em = parse_time(entry['end'])
                if eh is not None:
                    eh = f"{eh:02d}"
                    ws.cell(row=row, column=6, value=eh)
                if em is not None:
                    em = f"{em:02d}"
                    ws.cell(row=row, column=7, value=em)
            else:
                ws.cell(row=row, column=6, value="")
                ws.cell(row=row, column=7, value="")

            # 休憩時間は固定でJ列(10)に1を設定
            # break_timeが存在しない場合は0を設定
            if 'break_time' in entry:
                ws.cell(row=row, column=10, value=1)
                ws.cell(row=row, column=11, value="00")
            else:
                ws.cell(row=row, column=10, value=0)
                ws.cell(row=row, column=11, value="00")

    wb.save(excel_file)
    print(f"勤務ログをエクセルファイルに入力しました: {excel_file}")

def main():
    parser = argparse.ArgumentParser(description="勤務ログをエクセルファイルに自動入力するプログラム")
    parser.add_argument('--date', required=True, help='対象年月 (YYYYMM)')
    parser.add_argument('--file', required=True, help='エクセルファイルのパス')
    args = parser.parse_args()

    year_month = args.date
    excel_file = args.file

    # 勤務ログファイルのパスを作成
    script_dir = pathlib.Path(__file__).resolve().parent
    output_dir = script_dir / 'output'
    log_file_name = f"{year_month}_work_log.json"
    log_file_path = output_dir / log_file_name

    # 勤務ログを読み込む
    work_log = load_work_log(log_file_path)

    # エクセルファイルに勤務ログを自動入力する
    fill_work_log_to_excel(work_log, excel_file, year_month)

if __name__ == "__main__":
    main()